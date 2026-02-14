import pandas as pd
import openpyxl
import sys

# Ruta del archivo Excel
excel_path = r"C:\Users\PC-AORUS\Desktop\RFJ 2026\02_Core_Operativo\2. Core Operativo - Tejido Vivo y Creciente.xlsx"
output_path = r"C:\Users\PC-AORUS\Desktop\RFJ 2026\02_Core_Operativo\2. Core Operativo - Tejido Vivo y Creciente.md"

print("Iniciando conversión de Excel a Markdown...")
print(f"Archivo de entrada: {excel_path}")
print(f"Archivo de salida: {output_path}")

try:
    # Cargar el archivo Excel
    print("\nCargando archivo Excel...")
    wb = openpyxl.load_workbook(excel_path)
    print(f"✓ Excel cargado. Hojas encontradas: {wb.sheetnames}")
except Exception as e:
    print(f"✗ Error al cargar Excel: {e}")
    sys.exit(1)

# Abrir archivo de salida
try:
    with open(output_path, 'w', encoding='utf-8') as md_file:
        md_file.write("# Core Operativo - Tejido Vivo y Creciente\n\n")
        
        # Procesar cada hoja del Excel
        for sheet_name in wb.sheetnames:
            print(f"\nProcesando hoja: {sheet_name}")
            
            try:
                # Cargar la hoja con pandas
                df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
                
                # Eliminar filas y columnas completamente vacías
                df = df.dropna(how='all', axis=0)  # Eliminar filas vacías
                df = df.dropna(how='all', axis=1)  # Eliminar columnas vacías
                
                if df.empty:
                    print(f"  ⚠ Hoja vacía, omitiendo...")
                    continue
                
                # Rellenar valores NaN con cadena vacía
                df = df.fillna('')
                
                # Escribir el nombre de la hoja como título
                md_file.write(f"## {sheet_name}\n\n")
                
                # Convertir DataFrame a Markdown
                # Crear encabezado de tabla
                num_cols = len(df.columns)
                
                # Primera fila como encabezado
                header_row = "| " + " | ".join(str(val) for val in df.iloc[0]) + " |\n"
                separator = "|" + "|".join([" --- " for _ in range(num_cols)]) + "|\n"
                
                md_file.write(header_row)
                md_file.write(separator)
                
                # Escribir el resto de filas
                for idx in range(1, len(df)):
                    row = df.iloc[idx]
                    row_str = "| " + " | ".join(str(val) for val in row) + " |\n"
                    md_file.write(row_str)
                
                md_file.write("\n---\n\n")
                
                print(f"  ✓ Convertida ({len(df)} filas, {num_cols} columnas)")
            
            except Exception as e:
                print(f"  ✗ Error procesando hoja {sheet_name}: {e}")
                continue
    
    print(f"\n✓ Conversión completada exitosamente")
    print(f"✓ Archivo guardado en: {output_path}")

except Exception as e:
    print(f"✗ Error al escribir archivo Markdown: {e}")
    sys.exit(1)
